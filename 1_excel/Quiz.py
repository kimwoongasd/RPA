from openpyxl import Workbook

wb = Workbook()
ws = wb.active

ws.append(["학번", "출석", "퀴즈1", "퀴즈2", "중간고사", "기말고사", "프로잭트"])
scores = [(1,10,8,5,14,26,12),
(2,7,3,7,15,24,18),
(3,9,5,8,8,12,4),
(4,7,8,7,17,21,18),
(5,7,8,7,16,25,15),
(6,3,5,8,8,17,0),
(7,4,9,10,16,27,18),
(8,6,6,6,15,19,17),
(9,10,10,9,19,30,19),
(10,9,8,8,20,25,20)]

for s in scores:
    ws.append(s)

# for row in ws.rows:
#     for cell in row:
#         if cell.column == 4 and cell.row != 1:
#             cell.value = 10

for idx, cell in enumerate(ws["D"]):
    if idx == 0:
        continue
    cell.value = 10

ws["H1"] = '총점'
ws["I1"] = "성적"
# for i in range(2, 12):
#     ws[f"H{i}"] = f"=SUM(B{i}:G{i})"
# 총점과 성적 한번에 처리
for idx, score in enumerate(scores, start=2):
    sum_val = sum(score[1:]) - score[3] + 10
    ws.cell(row=idx, column=8).value = f"=SUM(B{idx}:G{idx})"
    
    if score[1] >= 5:
        if sum_val >= 90:
            ws[f"I{idx}"].value = "A"
        
        elif sum_val >= 80:
            ws[f"I{idx}"].value = "B"
            
        elif sum_val >= 70:
            ws[f"I{idx}"].value = "C"
            
        else:
            ws[f"I{idx}"].value = "D"
    else:
        ws[f"I{idx}"].value = "F"


# evaluate 되지 않은 상태의 데이터는 None이라고 출력
# for i in range(2, 12):
#     if ws[f"B{i}"].value >= 5:
#         if int(ws[f"H{i}"].value) >= 90:
#             ws[f"I{i}"] = 'A'
            
#         elif int(ws[f"H{i}"].value) >= 80:
#             ws[f"I{i}"] = 'B'
            
#         elif int(ws[f"H{i}"].value) >= 70:
#             ws[f"I{i}"] = 'C'
            
#         else:
#             ws[f"I{i}"] = 'D'
#     else:
#         ws[f"I{i}"] = 'F'

        
wb.save('scores.xlsx')
wb.close()