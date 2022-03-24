from openpyxl import Workbook

wb = Workbook()
ws = wb.active

# 병합하기
# b2:d2 까지 병합
ws.merge_cells("B2:D2")
ws["B2"].value = "Merge Cell"

wb.save("merge.xlsx")
wb.close()