from openpyxl import Workbook
from random import *
from openpyxl.utils.cell import coordinate_from_string

wb = Workbook()
ws = wb.active

# 1줄씩 데이터 넣기
ws.append(['번호', '영어', '수학'])

for i in range(1, 11):
    ws.append([i, randint(0, 100), randint(0, 100)])

# 영어 column만 가져오기
# col_b = ws["B"]
# for cell in col_b:
#     print(cell.value)

# 영어 수학 column 가져오기
# col_range = ws["B:C"]
# for cols in col_range:
#     for cell in cols:
#         print(cell.value)

# 1번쨰 row만 가져오기
# row_title = ws["1"]
# for i in row_title:
#     print(i.value)

# 2 ~ 6 번쨰 row만 가져오기
# rows_range = ws["2:6"]
# for i in rows_range:
#     for x in i:
#         # print(x.value, end=' ')
#         # print(x.coordinate, end=' ') # cell의 좌표정보를 가져온다
#         xy = coordinate_from_string(x.coordinate) # row와 column 를 분리
#         # print(xy, end=' ')
#         print(xy[0], end='')
#         print(xy[1], end=' ')
#     print()

# 전체 rows
# print(tuple(ws.rows))
# for row in tuple(ws.rows):
#     print(row[1].value)

# 다른 방법
# for row in ws.iter_rows():
#     print(row)


# 전체 columns
# print(tuple(ws.columns))
# for col in tuple(ws.columns):
#     print(col[0].value)

# 다른 방법
# for col in ws.iter_cols():
#     print(col)

# 2번쨰 row(가로)줄 부터 11번쨰 row까지, 2번쨰 column(열)부터 3번쨰 column까지
# for row in ws.iter_rows(min_row=2, max_row=11, min_col=2, max_col=3):
#     print(row[0].value, row[1].value) # 수학, 영어 점수
#     print(row)

for col in ws.iter_cols(min_row=1, max_row=5, min_col=1, max_col=3):
    print(col[3].value)
    

wb.save('sample2.xlsx')
wb.close()