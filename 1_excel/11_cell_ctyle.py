from openpyxl import load_workbook
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment

wb = load_workbook("sample2.xlsx")
ws = wb.active

a_1 = ws["A1"] # 번호
b_1 = ws["B1"] # 영어
c_1 = ws["C1"] # 수학

# 넓이 줄이기
ws.column_dimensions["A"].width = 5

# 높이 조절
ws.row_dimensions[1].height = 50

# 스타일 적용
# 색갈 : RGB 버전, 글꼴, 굵기
a_1.font = Font(color="FF0000", italic=True, bold=True)
# strike : 가운데 줄
b_1.font = Font(color="CC33FF", name="Arial", strike=True)
# underline : 밑줄
c_1.font = Font(color="0000FF", size=20, underline="single")

# 테두리
thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
a_1.border = thin_border
b_1.border = thin_border
c_1.border = thin_border

# 70점이상 셀을 초록색으로 적용
for row in ws.rows:
    for cell in row:
        # 중앙 정렬 center, left, right, top, bottom
        cell.alignment = Alignment(horizontal="center", vertical="center")
        
        if cell.column == 1:
            continue
        
        if isinstance(cell.value, int) and cell.value >= 70:
            cell.fill = PatternFill(fgColor="00FF00", fill_type='solid')
            cell.font = Font(color="FF0000")

# 틀 고정 b2기준
ws.freeze_panes = "B2"
      
wb.save("sample_style.xlsx")
wb.close()