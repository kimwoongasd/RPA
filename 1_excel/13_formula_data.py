from openpyxl import load_workbook

# wb = load_workbook("formula.xlsx")
# ws = wb.active

# # 수식 그래도 가져옴
# for row in ws.values:
#     for cell in row:
#         print(cell)


# 수식이 아닌 실제 데이터를 가져옴
wb = load_workbook("formula.xlsx", data_only=True)
ws = wb.active

# evaluate 되지 않은 상태의 데이터는 None이라고 출력
# 해결 방법 : 파일을 열고 저장하고 다시 실행
for row in ws.values:
    for cell in row:
        print(cell)