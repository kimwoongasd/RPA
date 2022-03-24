from openpyxl import load_workbook

wb = load_workbook('sample2.xlsx')
ws = wb.active

for row in ws.iter_rows(min_row=2):
    # 영어 점수
    if row[1].value > 50:
        print(row[0].value, "번 반 평균 이상")
        

for row in ws.iter_rows(max_row=1):
    for cell in row:
        if cell.value == '영어':
            cell.value = '컴퓨터'
            
wb.save('sample2_modified.xlsx')
wb.close()