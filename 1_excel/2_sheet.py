from tkinter import W
from openpyxl import Workbook

wb = Workbook()

# 새로운 sheet 기본 이름으로 생성
ws = wb.create_sheet()

# 타이틀 이름 변경
ws.title = 'Mysheet'

# RGB 형태로 값을 넣어주면 sheet 색상 변경
ws.sheet_properties.tabColor = 'ff66ff'

# 주어진 이름으로 sheet 생성
ws_1 = wb.create_sheet('Yoursheet')

# 2번쨰 index에 sheet 생성
ws_2 = wb.create_sheet("Newsheet", 2)

# dict 형태로 sheet에 접근
new_ws = wb["Newsheet"]

# 모든 sheet 이름 확인
print(wb.sheetnames)

# sheet 복사
new_ws["A1"] = 'Test'
target = wb.copy_worksheet(new_ws)
target.title = 'Copied sheet'

wb.save('sample.xlsx')
wb.close()
