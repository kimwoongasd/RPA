from openpyxl import load_workbook

wb = load_workbook("merge.xlsx")
ws = wb.active

# 병합 해제
ws.unmerge_cells("B2:D2")
ws["B2"] = "Unmerge Cell"

wb.save("unmerge.xlsx")
wb.close()